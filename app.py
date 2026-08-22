"""Streamlit SC Analysis tool, covering two models via a selector:

- Rental & Subscription (asset-lease variant): reproduces the SC Analysis
  for LTV and UE workbook for lease/rental contracts.
- Lending (loan-tape variant): reproduces the same tool's generic lending
  template, ported from the Lending/ project in this monorepo (itself
  derived from github.com/nephy-edge/screening-data-analysis-tool).

Shared infra (CA-bundle handling, the DeepInfra AI-suggestion call, the
mapping-cache read/write, the derived-column builder, the custom chart
builder, formatting) is unified; each model keeps its own schema, General
Inputs knobs, calc pipeline (scripts/rental_analysis vs
scripts/template_analysis), and tab rendering, since lease and loan
economics are different domains.
"""

import io
import json
import os
import re
import sys
import tempfile
from datetime import datetime as _dt

import altair as alt
import certifi
import pandas as pd
import requests
import streamlit as st
from openpyxl.chart import AreaChart, BarChart, LineChart, Reference, ScatterChart, Series

from dotenv import load_dotenv

load_dotenv()

from theme import inject_style, render_cover, render_masthead

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "scripts"))

from rental_analysis.general_inputs import GeneralInputs, DEFAULT_STATUS_MAP
from rental_analysis.data_questionnaire import QUESTIONS
from rental_analysis.data_input import (
    REQUIRED_COLUMNS, OPTIONAL_COLUMNS, apply_fallbacks, process_data_input,
)
from rental_analysis.asset_view import build_asset_view
from rental_analysis.repayment_curve import build_repayment_curve
from rental_analysis.lease_cohorts import build_cohorts
from rental_analysis.cohorts_for_x_or_more_loans import filter_cohorts
from rental_analysis.churn_analysis import ChurnAnalysis
from rental_analysis.ue_analysis import UeAnalysis
from rental_analysis.ltv_analysis import LtvAnalysis
from rental_analysis.ts_covenants import build_ts_covenants
from rental_analysis.general_analysis import describe as general_analysis

from template_analysis.general_inputs import GeneralInputs as LendingGeneralInputs
from template_analysis.data_questionnaire import QUESTIONS as LENDING_QUESTIONS
from template_analysis.data_input import process_data_input as lending_process_data_input
from template_analysis.cohorts import build_cohorts as lending_build_cohorts
from template_analysis.cohorts_for_x_or_more_loans import filter_cohorts as lending_filter_cohorts
from template_analysis.ltv_analysis import LtvAnalysis as LendingLtvAnalysis
from template_analysis.ue_analysis import UeAnalysis as LendingUeAnalysis
from template_analysis.general_analysis import describe as lending_general_analysis

DEEPINFRA_MODEL = "deepseek-ai/DeepSeek-V4-Flash-0731"
DEEPINFRA_CHAT_URL = "https://api.deepinfra.com/v1/openai/chat/completions"
_EXTRA_CA_PEM = os.path.join(os.path.dirname(__file__), "certs", "corporate_root.pem")


@st.cache_resource
def _ca_bundle_path() -> str:
    """Certifi's trust store plus this machine's corporate proxy root CA (if
    bundled), so HTTPS calls work on networks with TLS-inspecting proxies
    (e.g. Zscaler, Cisco Umbrella) without an env var set before launch."""
    if not os.path.exists(_EXTRA_CA_PEM):
        return certifi.where()
    with open(certifi.where(), "r", encoding="utf-8") as f:
        bundle = f.read()
    with open(_EXTRA_CA_PEM, "r", encoding="utf-8") as f:
        bundle += "\n" + f.read()
    tmp = tempfile.NamedTemporaryFile(mode="w", suffix=".pem", delete=False, encoding="utf-8")
    tmp.write(bundle)
    tmp.close()
    return tmp.name


def _get_slack_webhook_url():
    try:
        return st.secrets["SLACK_WEBHOOK_URL"]
    except Exception:
        return os.environ.get("SLACK_WEBHOOK_URL")


def _send_slack_feedback(message: str, model: str, user: str = "") -> tuple[bool, str]:
    """POST a feedback submission to the configured Slack Incoming Webhook."""
    webhook = _get_slack_webhook_url()
    if not webhook:
        return False, "SLACK_WEBHOOK_URL not configured."
    who = f" — {user.strip()}" if user.strip() else ""
    payload = {
        "text": (
            f"*Feedback* ({model}){who}\n{message.strip()}\n"
            f"_Submitted {_dt.now().strftime('%Y-%m-%d %H:%M')}_"
        )
    }
    try:
        resp = requests.post(
            webhook,
            json=payload,
            timeout=15,
            verify=_ca_bundle_path(),
        )
        resp.raise_for_status()
        return True, ""
    except Exception as e:
        return False, str(e)


# Read before set_page_config so the page title reflects the model chosen on
# a prior run (the selectbox widget itself is created further down).
is_lending = st.session_state.get("model_key", "rental") == "lending"

st.set_page_config(
    page_title="SC Analysis - Lending" if is_lending else "SC Analysis - LTV & UE",
    layout="wide",
    initial_sidebar_state="collapsed",
)
inject_style()
render_masthead("SC Analysis - Lending" if is_lending else "SC Analysis - Rental & Subscription")
render_cover(
    "Structured Credit Analysis - LTV & Unit Economics",
    "Upload loan-level portfolio data, map your columns to the Data Input "
    "template, and the analysis is computed automatically across all sheets."
    if is_lending else
    "Upload contract-level lease/rental data, map your columns to the Data Input "
    "template, and the analysis is computed automatically across all sheets.",
)

MODEL_LABELS = {"rental": "Rental & Subscription", "lending": "Lending"}
_default_model = st.session_state.get("model_key", "rental")
model_col, feedback_col = st.columns([5, 1], vertical_alignment="bottom")
with model_col:
    model_key = st.selectbox(
        "Select model", options=list(MODEL_LABELS.keys()),
        format_func=lambda k: MODEL_LABELS[k],
        index=list(MODEL_LABELS.keys()).index(_default_model) if _default_model in MODEL_LABELS else 0,
        key="model_key",
    )
if st.session_state.get("_active_model") != model_key:
    st.session_state["_active_model"] = model_key
    for k in ("uploaded_file_id", "analysis_ran", "analysis_mapping", "analysis_mapping_warnings",
              "analysis_gi_overrides", "derived_columns", "dc_ai_suggestion",
              "custom_chart_cards", "custom_chart_next_id", "export_charts"):
        st.session_state.pop(k, None)
is_lending = model_key == "lending"

with feedback_col:
    with st.popover("💬 Feedback"):
        st.caption("Share what worked or didn't work for you.")
        if not _get_slack_webhook_url():
            st.caption("Feedback routing not configured (SLACK_WEBHOOK_URL).")
        fb_user = st.text_input(
            "Your name or email (optional)",
            key="fb_user",
        )
        fb_text = st.text_area(
            "What worked / what didn't",
            key="fb_text",
            height=120,
        )
        if st.button("Send feedback", key="fb_send"):
            if not fb_text.strip():
                st.warning("Please enter some feedback first.")
            else:
                ok, err = _send_slack_feedback(
                    fb_text, MODEL_LABELS[model_key], fb_user
                )
                if ok:
                    st.session_state.pop("fb_text", None)
                    st.success("Thanks — feedback sent.")
                    st.rerun()
                else:
                    st.error(f"Could not send feedback: {err}")

RENTAL_CONFIG = dict(
    input_columns=[(c, True) for c in REQUIRED_COLUMNS] + [(c, False) for c in OPTIONAL_COLUMNS],
    date_fields={"start_date", "expected_end_date", "closed_date", "asset_recovery_date", "recovery_date"},
    numeric_fields={
        "downpayment", "total_contract_value", "monthly_expected_payment",
        "amount_expected_to_date", "total_paid", "cost_of_asset",
        "current_asset_value", "recovery_amount",
    },
    dayfirst=False,
    primary_date_field="start_date",
    mapping_cache_path=os.path.join(os.path.expanduser("~"), ".rental_sc_analysis_column_mappings.json"),
    needs_status_map=True,
    domain_hint="rental/lease contract-level dataset",
    derived_hint="If your file provides separate schedule fields but not amount_expected_to_date, "
                 "combine them here; the result becomes selectable in the mapping below.",
    derived_placeholder="For example, amount expected to date, calculated from monthly payment and months elapsed",
)
LENDING_CONFIG = dict(
    input_columns=[
        ("Loan ID", True), ("Disbursement Date", True), ("Expected Completion Date", True),
        ("Principal Value", True), ("Expected Interest", True), ("Expected Fee", False),
        ("Total Paid", True), ("Total Due", False),
    ],
    date_fields={"Disbursement Date", "Expected Completion Date"},
    numeric_fields={"Principal Value", "Expected Interest", "Expected Fee", "Total Paid", "Total Due"},
    dayfirst=True,
    primary_date_field="Disbursement Date",
    mapping_cache_path=os.path.join(os.path.expanduser("~"), ".sc_analysis_column_mappings.json"),
    needs_status_map=False,
    domain_hint="loan-portfolio spreadsheet",
    derived_hint="If your file provides Total GBV and Principal Value but not Expected Interest, "
                 "or separate Principal / Interest / Fee columns but not Total Due, combine them "
                 "here; the result becomes selectable in the mapping below.",
    derived_placeholder="For example, Expected Interest, calculated as Total GBV minus Principal Value",
)
active_cfg = LENDING_CONFIG if is_lending else RENTAL_CONFIG
INPUT_COLUMNS = active_cfg["input_columns"]
DATE_FIELDS = active_cfg["date_fields"]
NUMERIC_FIELDS = active_cfg["numeric_fields"]
REQUIRED_FIELDS = {t for t, required in INPUT_COLUMNS if required}
MAPPING_CACHE_PATH = active_cfg["mapping_cache_path"]
STATUS_CACHE_PATH = os.path.join(os.path.expanduser("~"), ".rental_sc_analysis_status_mappings.json")


def _cache_key(values) -> str:
    return "|".join(sorted(str(v) for v in values))


def _load_cache(path, key) -> dict | None:
    if not os.path.exists(path):
        return None
    try:
        with open(path, "r", encoding="utf-8") as f:
            cache = json.load(f)
    except (json.JSONDecodeError, OSError):
        return None
    return cache.get(key)


def _save_cache(path, key, value) -> None:
    cache = {}
    if os.path.exists(path):
        try:
            with open(path, "r", encoding="utf-8") as f:
                cache = json.load(f)
        except (json.JSONDecodeError, OSError):
            cache = {}
    cache[key] = value
    try:
        with open(path, "w", encoding="utf-8") as f:
            json.dump(cache, f, indent=2)
    except OSError:
        pass


def _validate_mapping(
    raw: pd.DataFrame, mapping: dict, required_fields, date_fields, numeric_fields, dayfirst=False
) -> tuple[list[str], list[str]]:
    errors, warnings = [], []
    for target, source in mapping.items():
        if not source:
            continue
        series = raw[source]
        non_null = series.notna()
        if non_null.sum() == 0:
            msg = f"**{target}** -> column '{source}' is entirely empty."
            (errors if target in required_fields else warnings).append(msg)
            continue
        if target in date_fields:
            parsed = _detect_date(series)
            fail_rate = 1 - (parsed.notna().sum() / non_null.sum())
            if fail_rate > 0.5:
                msg = f"**{target}** -> column '{source}' doesn't look like dates ({fail_rate:.0%} unparseable)."
                (errors if target in required_fields else warnings).append(msg)
        elif target in numeric_fields:
            parsed = pd.to_numeric(_clean_numeric(series), errors="coerce")
            fail_rate = 1 - (parsed.notna().sum() / non_null.sum())
            if fail_rate > 0.5:
                msg = f"**{target}** -> column '{source}' doesn't look numeric ({fail_rate:.0%} unparseable)."
                (errors if target in required_fields else warnings).append(msg)
    return errors, warnings


def fmt(value, spec="{:.1%}"):
    return spec.format(value) if pd.notna(value) else "n/a"


def _coerce_dates(raw: pd.DataFrame, date_fields, dayfirst=False) -> pd.DataFrame:
    for col in date_fields:
        if col in raw.columns:
            raw[col] = _detect_date(raw[col])
    return raw


# Canonical field -> acceptable header aliases, used to auto-normalise a file
# whose column names differ from the template's (strip spaces/punctuation/case
# before matching), so a "clean but differently-worded" upload still maps.
_FIELD_ALIASES = {
    "Loan ID": ["loan id", "loan_id", "account id", "account_id", "id", "loan no", "loan number", "contract no", "contract number"],
    "Disbursement Date": ["disbursement date", "disbursal date", "disbursed date", "funded date", "funding date", "start date", "origination date", "loan date", "issue date"],
    "Expected Completion Date": ["expected completion date", "maturity date", "expected end date", "end date", "expected maturity", "due date"],
    "Principal Value": ["principal value", "principal", "loan amount", "disbursement amount", "principal amount"],
    "Expected Interest": ["expected interest", "interest", "interest amount", "expected interest amount"],
    "Expected Fee": ["expected fee", "fee", "fee amount", "upfront fee", "expected fees"],
    "Total Paid": ["total paid", "amount paid", "total repayments", "payments", "total amount paid"],
    "Total Due": ["total due", "total dues calculated", "total due calculated", "pos", "outstanding", "balance", "amount due"],
    "Payment per Period": ["payment per period", "payment", "installment", "instalment", "monthly payment", "periodic payment"],
    "Payment Frequency": ["payment frequency", "frequency", "payment terms", "repayment frequency"],
    "contract_id": ["contract_id", "contract id", "contract no", "agreement id"],
    "asset_id": ["asset_id", "asset id", "asset no", "vin", "serial number"],
    "start_date": ["start_date", "start date", "contract start", "commencement date"],
    "status": ["status", "status label", "contract status", "stage", "status name"],
    "downpayment": ["downpayment", "down payment", "initial payment", "deposit"],
    "monthly_expected_payment": ["monthly_expected_payment", "monthly payment", "monthly rent", "monthly_rent", "expected monthly payment"],
    "total_paid": ["total_paid", "total paid", "amount paid"],
    "cost_of_asset": ["cost_of_asset", "cost of asset", "asset cost", "purchase price", "cost of the asset"],
    "expected_end_date": ["expected_end_date", "expected end date", "maturity date", "end date"],
    "closed_date": ["closed_date", "closed date", "close date", "cancellation date"],
    "asset_recovery_date": ["asset_recovery_date", "asset recovery date", "recovery date"],
    "total_contract_value": ["total_contract_value", "total contract value", "contract value"],
    "amount_expected_to_date": ["amount_expected_to_date", "amount expected to date", "expected to date"],
    "current_asset_value": ["current_asset_value", "current asset value", "asset value", "current value"],
    "recovery_date": ["recovery_date", "recovery date"],
    "recovery_amount": ["recovery_amount", "recovery amount"],
}

_ALL_FIELDS = set(_FIELD_ALIASES.keys()) | {
    "Begin Date", "Total Dues Calculated", "Delinquent Amount", "Write-off amount",
}


def _norm_key(value) -> str:
    return re.sub(r"[^a-z0-9]", "", str(value).lower())


def _clean_numeric(series) -> pd.Series:
    s = series.astype(str).str.strip()
    s = s.str.replace(r"[\$,£€\s%]", "", regex=True)
    s = s.str.replace(r"\(([^)]*)\)", r"-\1", regex=True)
    s = s.str.replace(",", "", regex=False)
    return s


def _detect_date(series) -> pd.Series:
    orig = series.copy()
    if pd.api.types.is_numeric_dtype(orig):
        # A date column can arrive as a raw Excel serial number (days since
        # 1899-12-30) when the source cell wasn't formatted as a date. Never
        # fall through to the string-parsing candidates below for numeric
        # input: pd.to_datetime treats a bare number as a nanosecond-epoch
        # timestamp and "succeeds" on every value, which would always win
        # the parse-rate comparison even though the result is silently wrong
        # (e.g. serial 44927 -> 1970-01-01, not 2023-01-01). Only trust
        # values in a plausible calendar-date range (~1900-2119).
        serial = pd.to_numeric(orig, errors="coerce")
        plausible = serial.where(serial.between(1, 80000))
        return pd.to_datetime(plausible, unit="D", origin="1899-12-30", errors="coerce")

    best, best_n = None, -1
    for kwargs in ({}, {"dayfirst": True}, {"yearfirst": True}):
        p = pd.to_datetime(orig, errors="coerce", **kwargs)
        n = int(p.notna().sum())
        if n > best_n:
            best_n, best = n, p
    return best if best_n > 0 else orig


def _normalize_columns(raw: pd.DataFrame, fields) -> pd.DataFrame:
    reverse = {}
    for canon in fields:
        reverse[_norm_key(canon)] = canon
        for alias in _FIELD_ALIASES.get(canon, []):
            reverse[_norm_key(alias)] = canon
    rename = {}
    for col in raw.columns:
        k = _norm_key(col)
        if k in reverse:
            canon = reverse[k]
            if canon not in raw.columns:
                rename[col] = canon
    return raw.rename(columns=rename)


def _format_normalize(raw: pd.DataFrame, date_fields, numeric_fields) -> pd.DataFrame:
    raw = _normalize_columns(raw, set(date_fields) | set(numeric_fields) | _ALL_FIELDS)
    for col in date_fields:
        if col in raw.columns and raw[col].dtype != "datetime64[ns]":
            raw[col] = _detect_date(raw[col])
    for col in numeric_fields:
        if col in raw.columns:
            cleaned = pd.to_numeric(_clean_numeric(raw[col]), errors="coerce")
            if cleaned.notna().sum() >= raw[col].notna().sum() * 0.5:
                raw[col] = cleaned
    return raw


DERIVED_OPS = {
    "+": lambda a, b: a + b,
    "-": lambda a, b: a - b,
    "×": lambda a, b: a * b,
    "÷": lambda a, b: a.divide(b).replace([float("inf"), float("-inf")], pd.NA),
}


def _apply_derived_columns(raw: pd.DataFrame, defs: list) -> pd.DataFrame:
    for d in defs:
        col_a = pd.to_numeric(raw[d["col_a"]], errors="coerce")
        col_b = pd.to_numeric(raw[d["col_b"]], errors="coerce")
        raw[d["name"]] = DERIVED_OPS[d["op"]](col_a, col_b)
    return raw


def _get_deepinfra_api_key():
    try:
        return st.secrets["DEEPINFRA_API_KEY"]
    except Exception:
        return os.environ.get("DEEPINFRA_API_KEY")


def _suggest_derived_column(user_request: str, columns: list, domain_hint: str) -> dict:
    """Ask DeepSeek V4 Flash (via DeepInfra) to turn a plain-English request into
    a two-column formula using only the columns actually present in the file."""
    api_key = _get_deepinfra_api_key()
    if not api_key:
        raise RuntimeError("No DEEPINFRA_API_KEY found. Add it to Streamlit secrets or the environment.")

    ops = list(DERIVED_OPS.keys())
    system_prompt = (
        f"You help map a {domain_hint}'s raw columns to a derived column. "
        "You can only combine exactly two existing columns with one of these operators: "
        f"{', '.join(ops)} (division). "
        "Pick the two columns and operator that best satisfy the user's request. "
        "If the request truly needs more than two columns or a non-arithmetic transform, "
        "still return your best two-column approximation and say so in the explanation.\n\n"
        f"Available columns (use these exact names): {', '.join(columns)}\n"
        f"Available operators (use exactly one of these characters): {', '.join(ops)}\n\n"
        "Respond with ONLY a JSON object, no markdown fences, matching this shape:\n"
        '{"name": "<short column name>", "col_a": "<one of the available columns>", '
        '"op": "<one of the available operators>", "col_b": "<one of the available columns>", '
        '"explanation": "<one sentence>"}'
    )

    resp = requests.post(
        DEEPINFRA_CHAT_URL,
        headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
        json={
            "model": DEEPINFRA_MODEL,
            "messages": [
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_request},
            ],
            "response_format": {"type": "json_object"},
            "max_tokens": 512,
        },
        timeout=30,
        verify=_ca_bundle_path(),
    )
    resp.raise_for_status()
    content = resp.json()["choices"][0]["message"]["content"]
    suggestion = json.loads(content)

    missing = [k for k in ("name", "col_a", "op", "col_b", "explanation") if k not in suggestion]
    if missing:
        raise ValueError(f"Model response missing fields: {', '.join(missing)}")
    if suggestion["col_a"] not in columns or suggestion["col_b"] not in columns:
        raise ValueError("Model suggested a column that isn't in your file.")
    if suggestion["op"] not in DERIVED_OPS:
        raise ValueError(f"Model suggested an unsupported operator: {suggestion['op']}")
    return suggestion


# Short human-readable meaning for each Data Input template field, used to prompt
# the auto-fill model so it maps raw file columns to the right template fields.
FIELD_DESCRIPTIONS = {
    # Lending
    "Loan ID": "unique loan/account identifier",
    "Disbursement Date": "date the loan was disbursed/funded",
    "Expected Completion Date": "expected maturity or completion date",
    "Principal Value": "principal amount lent",
    "Expected Interest": "expected interest amount",
    "Expected Fee": "expected fee amount",
    "Total Paid": "total amount the borrower has paid",
    "Total Due": "total amount owed (principal + interest + fees)",
    # Rental & Subscription
    "contract_id": "unique contract/lease identifier",
    "asset_id": "asset/device identifier",
    "start_date": "date the contract started",
    "status": "contract status label",
    "downpayment": "initial down-payment amount",
    "monthly_expected_payment": "expected monthly payment amount",
    "total_paid": "total amount paid to date",
    "cost_of_asset": "cost of the asset",
    "expected_end_date": "expected end/maturity date",
    "closed_date": "date the contract was closed",
    "asset_recovery_date": "date the asset was recovered",
    "total_contract_value": "total contract value",
    "amount_expected_to_date": "amount expected to have been paid to date",
    "current_asset_value": "current value of the asset",
    "recovery_date": "date of recovery",
    "recovery_amount": "recovery amount",
}


def _suggest_mapping(columns: list, input_columns: list) -> dict:
    """Ask DeepSeek V4 Flash (via DeepInfra) to guess the best mapping of the
    uploaded file's raw columns to each Data Input template field for the active
    model. Returns {target_field: source_column} using only columns actually
    present; fields with no sensible source are left unmapped (None)."""
    api_key = _get_deepinfra_api_key()
    if not api_key:
        return {}

    fields = "\n".join(
        f"- {target} ({'required' if required else 'optional'}): {FIELD_DESCRIPTIONS.get(target, '')}"
        for target, required in input_columns
    )
    system_prompt = (
        "You are mapping a model's raw column names to a fixed set of template "
        "fields. For each template field below, choose the single raw column "
        "that best matches its meaning.\n\n"
        f"Available raw columns (use these exact names, or null if none fit): "
        f"{', '.join(columns)}\n\n"
        "Template fields to fill:\n"
        f"{fields}\n\n"
        "Respond with ONLY a JSON object, no markdown fences, mapping each "
        "template field to a raw column name (exact string) or null when no "
        "column plausibly matches. Example:\n"
        '{"Loan ID": "account_id", "Disbursement Date": "funded_at", '
        '"Total Due": null}'
    )

    resp = requests.post(
        DEEPINFRA_CHAT_URL,
        headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
        json={
            "model": DEEPINFRA_MODEL,
            "messages": [
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": "Guess the best mapping for these columns."},
            ],
            "response_format": {"type": "json_object"},
            "max_tokens": 512,
        },
        timeout=30,
        verify=_ca_bundle_path(),
    )
    resp.raise_for_status()
    content = resp.json()["choices"][0]["message"]["content"]
    suggestions = json.loads(content)

    valid_columns = set(str(c) for c in columns)
    mapping = {}
    for target, _ in input_columns:
        src = suggestions.get(target)
        if src is None:
            mapping[target] = None
            continue
        src = str(src).strip()
        mapping[target] = src if src in valid_columns else None
    return mapping


def _add_reference_line(chart, value, label, color="#d62728", x_anchor=None):
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return chart
    ref = pd.DataFrame({"v": [value], "label": [label]})
    rule = alt.Chart(ref).mark_rule(color=color, strokeDash=[6, 6]).encode(y=alt.Y("v:Q"))
    if x_anchor is not None:
        ref["xa"] = [x_anchor]
        text = alt.Chart(ref).mark_text(
            color=color, dx=-6, dy=-6, align="right", fontSize=11, fontWeight="bold"
        ).encode(x="xa:T", y=alt.Y("v:Q"), text="label:N")
    else:
        text = alt.Chart(ref).mark_text(
            color=color, dy=-7, dx=6, align="left", fontSize=11, fontWeight="bold"
        ).encode(x=alt.value(70), y=alt.Y("v:Q"), text="label:N")
    return chart + rule + text


def _render_custom_visualizations_tab(data_sources: dict):
    """Generic ad-hoc chart builder shared by both models. `data_sources` maps
    a display name to the DataFrame it plots from for the active model."""
    st.subheader("Custom Visualizations")
    st.caption("Build your own charts from the mapped and computed data. Add as many chart cards as you like.")

    if "custom_chart_cards" not in st.session_state:
        st.session_state["custom_chart_cards"] = [0]
        st.session_state["custom_chart_next_id"] = 1

    def _render_chart_card(card_id):
        k = lambda name: f"cv_{name}_{card_id}"
        cv1, cv2 = st.columns([1, 1])
        with cv1:
            source_name = st.selectbox("Data source", options=list(data_sources.keys()), key=k("source"))
        source_df = data_sources[source_name]
        if source_df is None or source_df.empty:
            st.info("This data source has no rows to plot.")
            return

        all_cols = list(source_df.columns)
        numeric_cols = [c for c in all_cols if pd.api.types.is_numeric_dtype(source_df[c])]
        datetime_cols = [c for c in all_cols if pd.api.types.is_datetime64_any_dtype(source_df[c])]
        categorical_cols = [c for c in all_cols if c not in numeric_cols and c not in datetime_cols]

        # Smart defaults for a fresh chart card: an origination-date-like column on X
        # and a principal-like column on Y, summed - so it opens showing something
        # meaningful (e.g. Principal by Cohort) instead of an arbitrary first column.
        # Only affects the initial selectbox index; once a user picks a value for this
        # card, Streamlit remembers it under that widget's key on reruns.
        def _pick(cols, keywords, fallback):
            for kw in keywords:
                for c in cols:
                    if kw in c.lower():
                        return c
            return fallback

        default_x = _pick(
            datetime_cols, ["cohort", "disbursement", "origination", "start"],
            datetime_cols[0] if datetime_cols else all_cols[0],
        )
        y_options = numeric_cols if numeric_cols else all_cols
        default_y = _pick(numeric_cols, ["principal"], y_options[0])
        default_agg = "Sum" if default_x in datetime_cols and default_y in numeric_cols else "(none - raw rows)"
        chart_options = ["Line", "Bar", "Scatter", "Area"]
        default_kind = "Bar" if default_agg == "Sum" else "Line"

        with cv2:
            chart_kind = st.selectbox("Chart type", options=chart_options, index=chart_options.index(default_kind), key=k("kind"))

        cv3, cv4, cv5, cv6 = st.columns(4)
        with cv3:
            x_col = st.selectbox("X axis", options=all_cols, index=all_cols.index(default_x), key=k("x"))
        with cv4:
            y_col = st.selectbox("Y axis", options=y_options, index=y_options.index(default_y), key=k("y"))
        with cv5:
            color_col = st.selectbox("Group / color by (optional)", options=["(none)"] + [c for c in categorical_cols if c != x_col], key=k("color"))
        with cv6:
            agg_options = ["(none - raw rows)", "Sum", "Mean", "Count", "Median", "Min", "Max"]
            agg_func = st.selectbox("Aggregate Y by X (optional)", options=agg_options, index=agg_options.index(default_agg), key=k("agg"))

        base_cols = [c for c in {x_col, y_col, color_col} if c in source_df.columns]
        plot_df = source_df[base_cols].dropna(subset=[x_col, y_col])

        if agg_func != "(none - raw rows)" and not plot_df.empty:
            group_cols = [x_col] + ([color_col] if color_col != "(none)" else [])
            agg_name_map = {"Sum": "sum", "Mean": "mean", "Count": "count", "Median": "median", "Min": "min", "Max": "max"}
            if agg_func == "Count":
                plot_df = plot_df.groupby(group_cols, dropna=False)[y_col].count().reset_index()
            else:
                plot_df = plot_df.groupby(group_cols, dropna=False)[y_col].agg(agg_name_map[agg_func]).reset_index()

        if plot_df.empty:
            st.info("No rows with data for the selected columns.")
            return

        x_type = "T" if x_col in datetime_cols else ("O" if x_col in categorical_cols else "Q")
        mark_map = {
            "Line": alt.Chart(plot_df).mark_line(point=True),
            "Bar": alt.Chart(plot_df).mark_bar(),
            "Scatter": alt.Chart(plot_df).mark_circle(size=60),
            "Area": alt.Chart(plot_df).mark_area(opacity=0.6),
        }
        base = mark_map[chart_kind]
        y_title = f"{agg_func} of {y_col}" if agg_func != "(none - raw rows)" else y_col
        chart_title = f"{y_title} by {x_col}" + (f" ({color_col})" if color_col != "(none)" else "")

        encode_kwargs = dict(
            x=alt.X(f"{x_col}:{x_type}", title=x_col),
            y=alt.Y(f"{y_col}:Q", title=y_title),
            tooltip=[x_col, y_col] + ([color_col] if color_col != "(none)" else []),
        )
        if color_col != "(none)":
            encode_kwargs["color"] = alt.Color(f"{color_col}:N", title=color_col)

        custom_chart = base.encode(**encode_kwargs).properties(title=chart_title, height=400)
        st.altair_chart(custom_chart, width="stretch")

        bc1, bc2 = st.columns([1, 1])
        with bc1:
            if st.button("Add this chart to the Excel export", key=k("add_export")):
                export_charts = st.session_state.setdefault("export_charts", [])
                export_charts.append({
                    "title": chart_title, "kind": chart_kind, "x_col": x_col, "y_col": y_col,
                    "color_col": None if color_col == "(none)" else color_col, "y_title": y_title,
                    "data": plot_df[[c for c in {x_col, y_col, color_col} if c in plot_df.columns]].copy(),
                })
                st.success(f"Added '{chart_title}' to the Excel export queue.")
        with bc2:
            if len(st.session_state["custom_chart_cards"]) > 1:
                if st.button("Remove this chart card", key=k("remove_card")):
                    st.session_state["custom_chart_cards"].remove(card_id)
                    st.rerun()

    for i, card_id in enumerate(st.session_state["custom_chart_cards"]):
        st.markdown(f"#### Chart {i + 1}")
        _render_chart_card(card_id)
        st.markdown("---")

    if st.button("Add another chart", key="cv_add_card"):
        st.session_state["custom_chart_cards"].append(st.session_state["custom_chart_next_id"])
        st.session_state["custom_chart_next_id"] += 1
        st.rerun()

    export_charts = st.session_state.get("export_charts", [])
    if export_charts:
        st.markdown("**Charts queued for the Excel export:**")
        for i, ec in enumerate(export_charts):
            qc1, qc2 = st.columns([5, 1])
            qc1.write(f"{i + 1}. {ec['title']} ({ec['kind']})")
            if qc2.button("Remove", key=f"cv_remove_{i}"):
                export_charts.pop(i)
                st.rerun()
        if st.button("Clear all queued charts", key="cv_clear_export"):
            st.session_state["export_charts"] = []
            st.rerun()


def _write_custom_charts_sheet(writer, export_charts):
    """Shared by both export functions - identical chart-embedding logic,
    independent of which model produced the queued chart definitions."""
    if not export_charts:
        return
    cc_ws = writer.book.create_sheet("Custom Charts")
    row = 1
    for ec in export_charts:
        if ec["color_col"]:
            wide = ec["data"].pivot_table(index=ec["x_col"], columns=ec["color_col"], values=ec["y_col"], aggfunc="first").reset_index()
        else:
            wide = ec["data"][[ec["x_col"], ec["y_col"]]]
        header_row = row
        wide.to_excel(writer, sheet_name="Custom Charts", startrow=row - 1, index=False)
        data_start = header_row + 1
        data_end = header_row + len(wide)
        n_series_cols = len(wide.columns) - 1
        chart_map = {"Line": LineChart, "Bar": BarChart, "Area": AreaChart}
        if ec["kind"] in chart_map:
            c = chart_map[ec["kind"]]()
            c.title = ec["title"]
            c.height, c.width = 10, 20
            data = Reference(cc_ws, min_col=2, min_row=header_row, max_col=1 + n_series_cols, max_row=data_end)
            cats = Reference(cc_ws, min_col=1, min_row=data_start, max_row=data_end)
            c.add_data(data, titles_from_data=True)
            c.set_categories(cats)
            cc_ws.add_chart(c, f"{chr(ord('A') + n_series_cols + 3)}{header_row}")
        else:
            sc = ScatterChart()
            sc.title = ec["title"]
            sc.height, sc.width = 10, 20
            sc.x_axis.title = ec["x_col"]
            sc.y_axis.title = ec["y_title"]
            xvalues = Reference(cc_ws, min_col=1, min_row=data_start, max_row=data_end)
            for col_idx in range(2, 2 + n_series_cols):
                yvalues = Reference(cc_ws, min_col=col_idx, min_row=header_row, max_row=data_end)
                series = Series(yvalues, xvalues, title_from_data=True)
                series.marker.symbol = "circle"
                series.graphicalProperties.line.noFill = True
                sc.series.append(series)
            cc_ws.add_chart(sc, f"{chr(ord('A') + n_series_cols + 3)}{header_row}")
        row = data_end + 3


def _read_tabular_file(uploaded) -> pd.DataFrame:
    """Read an uploaded CSV/XLSX, robust to a workbook with more than one
    sheet (defaults to reading the first, but lets the user pick) and to a
    header row that isn't row 0 - e.g. a title/banner row above the real
    column headers, which would otherwise silently produce "Unnamed: N"
    columns instead of an error."""
    if uploaded.name.endswith(".csv"):
        return pd.read_csv(uploaded)

    xls = pd.ExcelFile(uploaded)
    sheet_name = xls.sheet_names[0]
    if len(xls.sheet_names) > 1:
        sheet_name = st.selectbox(
            "This file has multiple sheets - which one has your loan-level data?",
            options=xls.sheet_names, key="upload_sheet_name",
        )

    df = xls.parse(sheet_name)
    unnamed_frac = sum(str(c).startswith("Unnamed:") for c in df.columns) / max(len(df.columns), 1)
    if unnamed_frac >= 0.5:
        preview = xls.parse(sheet_name, header=None, nrows=10)
        for i in range(1, len(preview)):
            row = preview.iloc[i]
            if row.notna().mean() > 0.7 and row.dropna().map(lambda v: isinstance(v, str)).mean() > 0.7:
                df = xls.parse(sheet_name, header=i)
                break
    return df


uploaded = st.file_uploader(
    "Choose a CSV or Excel file", type=["csv", "xlsx"], key=f"uploader_{model_key}"
)

if not uploaded:
    st.info("Upload a contract-level file to begin.")
    st.stop()

if st.session_state.get("uploaded_file_id") != uploaded.file_id:
    st.session_state["uploaded_file_id"] = uploaded.file_id
    st.session_state["analysis_ran"] = False
    st.session_state.pop("analysis_mapping", None)
    st.session_state.pop("analysis_gi_overrides", None)
    st.session_state["derived_columns"] = []
    for target, _ in INPUT_COLUMNS:
        st.session_state.pop(f"map_{target}", None)

raw = _read_tabular_file(uploaded)
raw = _format_normalize(raw, DATE_FIELDS, NUMERIC_FIELDS)

st.success(f"Loaded {len(raw):,} rows - {len(raw.columns)} columns.")
st.dataframe(raw.head(10), width="stretch", height=300)
st.caption("Columns in your file: " + ", ".join(map(str, raw.columns)))

if "derived_columns" not in st.session_state:
    st.session_state["derived_columns"] = []

with st.expander("Derive a missing column from existing fields"):
    st.caption(active_cfg["derived_hint"])
    st.markdown("**Describe the calculation and let AI suggest the formula.**")
    ai1, ai2 = st.columns([4, 1])
    with ai1:
        ai_request = st.text_area(
            "Describe the calculation",
            placeholder=active_cfg["derived_placeholder"],
            key="dc_ai_request", height=70,
        )
    with ai2:
        st.write("")
        ask_ai = st.button("Suggest formula", key="dc_ai_ask")

    if ask_ai:
        if not ai_request.strip():
            st.warning("Describe the calculation before requesting a suggestion.")
        else:
            try:
                with st.spinner("Asking DeepSeek..."):
                    st.session_state["dc_ai_suggestion"] = _suggest_derived_column(
                        ai_request, list(raw.columns), active_cfg["domain_hint"]
                    )
            except Exception as e:
                st.session_state["dc_ai_suggestion"] = None
                st.error(f"Unable to generate a suggestion: {e}")

    suggestion = st.session_state.get("dc_ai_suggestion")
    if suggestion:
        st.info(
            f"**Suggested:** {suggestion['name']} = {suggestion['col_a']} "
            f"{suggestion['op']} {suggestion['col_b']}\n\n{suggestion['explanation']}"
        )
        if st.button("Use this suggestion", key="dc_ai_use"):
            name = suggestion["name"].strip()
            existing_names = {d["name"] for d in st.session_state["derived_columns"]}
            if name in raw.columns or name in existing_names:
                base, i = name, 2
                while f"{base} ({i})" in raw.columns or f"{base} ({i})" in existing_names:
                    i += 1
                name = f"{base} ({i})"
            st.session_state["derived_columns"].append({
                "name": name, "col_a": suggestion["col_a"], "op": suggestion["op"], "col_b": suggestion["col_b"],
            })
            st.session_state["dc_ai_suggestion"] = None
            st.session_state["dc_ai_request"] = ""
            st.rerun()

    st.markdown("**Or build it manually:**")
    dc1, dc2, dc3, dc4 = st.columns([2, 2, 1, 2])
    with dc1:
        new_name = st.text_input("New column name", key="dc_name")
    with dc2:
        col_a = st.selectbox("Column A", options=list(raw.columns), key="dc_col_a")
    with dc3:
        op = st.selectbox("Operator", options=list(DERIVED_OPS.keys()), key="dc_op")
    with dc4:
        col_b = st.selectbox("Column B", options=list(raw.columns), key="dc_col_b")

    if st.button("Add derived column", key="dc_add"):
        existing_names = {d["name"] for d in st.session_state["derived_columns"]}
        if not new_name.strip():
            st.error("Give the derived column a name.")
        elif new_name in raw.columns or new_name in existing_names:
            st.error(f"'{new_name}' already exists - choose a different name.")
        else:
            st.session_state["derived_columns"].append({"name": new_name.strip(), "col_a": col_a, "op": op, "col_b": col_b})
            st.rerun()

    if st.session_state["derived_columns"]:
        st.markdown("**Derived columns:**")
        for i, d in enumerate(st.session_state["derived_columns"]):
            rc1, rc2 = st.columns([5, 1])
            rc1.write(f"`{d['name']}` = {d['col_a']} {d['op']} {d['col_b']}")
            if rc2.button("Remove", key=f"dc_remove_{i}"):
                st.session_state["derived_columns"].pop(i)
                st.rerun()

raw = _apply_derived_columns(raw, st.session_state["derived_columns"])
if st.session_state["derived_columns"]:
    st.dataframe(raw[[d["name"] for d in st.session_state["derived_columns"]]].head(10), width="stretch", height=150)

cached_mapping = _load_cache(MAPPING_CACHE_PATH, _cache_key(raw.columns)) or {}
if cached_mapping:
    st.caption("A saved mapping was found for a file with these same column headers - pre-filled below.")

# AI auto-fill: guess the best column mapping for this file's headers (per model).
guess_key = f"ai_mapping_guess_{model_key}_{_cache_key(raw.columns)}"
if guess_key not in st.session_state:
    if _get_deepinfra_api_key():
        try:
            with st.spinner("Asking the AI to auto-fill the column mapping..."):
                st.session_state[guess_key] = _suggest_mapping(list(raw.columns), INPUT_COLUMNS)
        except Exception as e:
            st.session_state[guess_key] = {}
            st.warning(f"AI auto-fill wasn't available ({e}). You can still map columns manually.")
    else:
        st.session_state[guess_key] = {}
ai_guess = st.session_state.get(guess_key) or {}

with st.form("column_mapping"):
    st.subheader("Map your columns to the Data Input template")
    st.caption("For each template field below, select the matching column in your file. Required fields must be mapped to run the analysis.")
    used = set()
    mapping = {}
    for target, required in INPUT_COLUMNS:
        options = ["(not provided)"] + [c for c in raw.columns if c not in used]
        # Precedence: saved mapping > AI guess > "(not provided)".
        default_choice = cached_mapping.get(target) or ai_guess.get(target)
        default_index = options.index(default_choice) if default_choice in options else 0
        chosen = st.selectbox(
            f"Map to **{target}** ({'required' if required else 'optional'})",
            options=options, index=default_index, key=f"map_{target}",
        )
        mapping[target] = None if chosen == "(not provided)" else chosen
        if chosen != "(not provided)":
            used.add(chosen)

    # Date of extraction defaults to the max of whichever raw column the user
    # just mapped to the primary date field (Disbursement Date / start_date) -
    # not a guess at the raw header's name, since that column could be called
    # anything before mapping renames it.
    default_extraction = pd.Timestamp.now().normalize()
    primary_col = mapping.get(active_cfg["primary_date_field"])
    if primary_col:
        parsed_primary = pd.to_datetime(raw[primary_col], errors="coerce")
        if parsed_primary.notna().any():
            default_extraction = parsed_primary.max()

    st.subheader("General Inputs")
    if is_lending:
        gc1, gc2, gc3 = st.columns(3)
        with gc1:
            extraction_date = st.date_input(
                "Date of extraction", value=default_extraction.date(),
                help="The max date in the loan tape. Used to identify which loans have reached their maturity.",
            )
        with gc2:
            days_after_term = st.number_input(
                "Days after term", value=90, min_value=0,
                help="Days after term used to compute loss rate (default 90 = Term + 3 months). "
                     "Modify only if the company has significant repayments after 3 months from term.",
            )
        with gc3:
            min_loans_per_cohort = st.number_input(
                "Minimum loans per cohort", value=10, min_value=0,
                help="Affects the cohort stressed loss rate: requires a minimum number of observations to "
                     "include a cohort in the cohort loss rate distribution (default 10).",
            )
    else:
        gc1, gc2, gc3 = st.columns(3)
        with gc1:
            extraction_date = st.date_input(
                "Date of extraction", value=default_extraction.date(),
                help="The max date in the loan tape. Used to identify which loans have reached their maturity.",
            )
            days_after_term = st.number_input(
                "Days after term", value=0, min_value=0,
                help="Days after term used to compute loss rate (default 90 = Term + 3 months). "
                     "Modify only if the company has significant repayments after 3 months from term.",
            )
        with gc2:
            months_since_default = st.number_input("Months since default", value=3, min_value=0)
            min_loans_per_cohort = st.number_input(
                "Minimum loans per cohort", value=20, min_value=0,
                help="Affects the cohort stressed loss rate: requires a minimum number of observations to "
                     "include a cohort in the cohort loss rate distribution (default 10).",
            )
        with gc3:
            useful_life_years = st.number_input("Useful life of asset (years)", value=3.0, min_value=0.1)

        st.subheader("Lendable status labels")
        lc1, lc2, lc3 = st.columns(3)
        with lc1:
            open_label = st.text_input("Status used for active", value="Open")
        with lc2:
            closed_label = st.text_input("Status used for canceled", value="Closed")
        with lc3:
            paidoff_label = st.text_input("Status used for paid-off", value="Paid-off")

    submitted = st.form_submit_button("Run analysis")

if submitted:
    missing_required = [t for t, required in INPUT_COLUMNS if required and not mapping[t]]
    if missing_required:
        st.error("Map these required fields before running: " + ", ".join(missing_required))
        st.stop()
    mapping_errors, mapping_warnings = _validate_mapping(
        raw, mapping, REQUIRED_FIELDS, DATE_FIELDS, NUMERIC_FIELDS, active_cfg["dayfirst"]
    )
    if mapping_errors:
        st.error("Fix these mappings before running:\n\n" + "\n".join(f"- {e}" for e in mapping_errors))
        st.stop()
    st.session_state["analysis_mapping"] = mapping
    st.session_state["analysis_mapping_warnings"] = mapping_warnings
    # Date of extraction follows the workbook formula =MAX(primary date column)
    # (General Inputs C3 in the sheets), overriding the calendar default once
    # the primary-date-field mapping is known.
    analysis_extraction = pd.Timestamp(extraction_date)
    start_src = mapping.get(active_cfg["primary_date_field"])
    if start_src:
        parsed_start = _detect_date(raw[start_src])
        if parsed_start.notna().any():
            analysis_extraction = pd.Timestamp(parsed_start.max())
    overrides = {
        "extraction_date": analysis_extraction,
        "days_after_term": days_after_term,
        "min_loans_per_cohort": min_loans_per_cohort,
    }
    if not is_lending:
        overrides.update({
            "months_since_default": months_since_default,
            "useful_life_years": useful_life_years,
            "open_label": open_label,
            "closed_label": closed_label,
            "paidoff_label": paidoff_label,
        })
    st.session_state["analysis_gi_overrides"] = overrides
    st.session_state["analysis_ran"] = True
    _save_cache(MAPPING_CACHE_PATH, _cache_key(raw.columns), mapping)

if not st.session_state.get("analysis_ran"):
    st.info("Map your file's columns above, then click 'Run analysis' to compute all sheets.")
    st.stop()

mapping = st.session_state["analysis_mapping"]
rename_map = {src: tgt for tgt, src in mapping.items() if src}
raw = raw.rename(columns=rename_map)
_coerce_dates(raw, DATE_FIELDS, active_cfg["dayfirst"])
# _format_normalize() coerced numeric_fields before this rename, so it only ever
# caught columns whose raw header already auto-aliased to a template name.
# Anything the user had to manually map above skipped coercion entirely and
# stayed as raw text - re-run numeric coercion now that the mapping is final.
for col in NUMERIC_FIELDS:
    if col in raw.columns and raw[col].dtype != "float64":
        raw[col] = pd.to_numeric(_clean_numeric(raw[col]), errors="coerce")

if not is_lending:
    # The workbooks name the asset recovery sale/proceeds column "recovery_value",
    # while the template/calculation field is "recovery_amount". Auto-alias so a
    # workbook-named file still feeds recoveries into the recovery-based metrics
    # (% collected on principal, LTV recovery) without the user having to map it.
    if "recovery_amount" not in raw.columns and "recovery_value" in raw.columns:
        raw["recovery_amount"] = raw["recovery_value"]

unmapped = [t for t, _ in INPUT_COLUMNS if t not in raw.columns]
if unmapped:
    st.warning("Not mapped (optional) - related metrics will be unavailable: " + ", ".join(unmapped))

mapping_warnings = st.session_state.get("analysis_mapping_warnings") or []
if mapping_warnings:
    st.warning("Mapping quality warnings:\n\n" + "\n".join(f"- {w}" for w in mapping_warnings))

if active_cfg["needs_status_map"]:
    raw_statuses = sorted(raw["status"].dropna().unique().tolist())
    status_key = _cache_key(raw_statuses)
    cached_status_map = _load_cache(STATUS_CACHE_PATH, status_key) or {}
    seed = [
        {"Client status": s, "Lendable status": cached_status_map.get(s, DEFAULT_STATUS_MAP.get(s, ""))}
        for s in raw_statuses
    ]
    st.subheader("Status mapping")
    st.caption("Map each raw status value in your file to a standardised Lendable status.")
    status_map_df = st.data_editor(pd.DataFrame(seed), hide_index=True, width="stretch", key="status_map_editor")
    status_map = dict(zip(status_map_df["Client status"], status_map_df["Lendable status"]))

    unmapped_statuses = [s for s, v in status_map.items() if not v]
    if unmapped_statuses:
        st.warning("Unmapped status value(s), fill these in above: " + ", ".join(unmapped_statuses))
        st.stop()
    _save_cache(STATUS_CACHE_PATH, status_key, status_map)
else:
    status_map = None

gi_overrides = dict(st.session_state["analysis_gi_overrides"])
if is_lending:
    gi = LendingGeneralInputs(raw, **gi_overrides)
else:
    gi_overrides["status_map"] = status_map
    gi = GeneralInputs(raw, **gi_overrides)

with st.spinner("Running analysis..."):
    if is_lending:
        df = lending_process_data_input(raw, gi.extraction_date, gi.days_after_term)
        cohorts = lending_build_cohorts(df)
        filtered = lending_filter_cohorts(cohorts, gi.min_loans_per_cohort)
        st.caption(f"Mapped & computed columns: {', '.join(df.columns)}")
        if "Total Due" not in df.columns:
            st.warning("No 'Total Due' or payment schedule mapped - loss rate proxy and PvD ratio will be unavailable.")
        ltv = LendingLtvAnalysis(df, filtered)
        ue = LendingUeAnalysis(df)
        ue_data, ltv_data = ue.as_dict(), ltv.as_dict()
        lending_chart_data = cohorts.dropna(subset=["Loss Rate"]).copy()
        if not lending_chart_data.empty:
            lending_chart_data["Fee %"] = lending_chart_data["Total Fee"] / lending_chart_data["Total Principal"]
            lending_chart_data["Interest %"] = lending_chart_data["Total Interest"] / lending_chart_data["Total Principal"]
    else:
        fallback_df, fallback_notes = apply_fallbacks(raw, gi.useful_life_years, gi.extraction_date)
        df = process_data_input(fallback_df, gi.as_calc_dict())
        av = build_asset_view(df, gi.as_calc_dict())
        curve = build_repayment_curve(av, gi.as_calc_dict())
        cohorts = build_cohorts(df, gi.as_calc_dict())
        filtered = filter_cohorts(cohorts, gi.min_loans_per_cohort)
        churn = ChurnAnalysis(cohorts, gi.as_calc_dict())
        ue = UeAnalysis(df, av, curve, gi.as_calc_dict())
        ltv = LtvAnalysis(df, churn, gi.as_calc_dict())
        ts = build_ts_covenants(ue.as_dict(), ltv.as_dict(), curve)
        st.caption(f"Mapped & computed columns: {', '.join(df.columns)}")
        ue_data, ltv_data, ca_data = ue.as_dict(), ltv.as_dict(), churn.as_dict()

if not is_lending and fallback_notes:
    with st.expander(f"Data quality: {len(fallback_notes)} fallback rule(s) applied"):
        for note in fallback_notes:
            st.write("-", note)

active_questions = LENDING_QUESTIONS if is_lending else QUESTIONS

if is_lending:
    tab_names = [
        "General Inputs", "Data Questionnaire", "Data Input",
        "Cohorts", "Cohorts for X or more loans",
        "LTV Analysis", "Unit Economics Analysis", "General Analysis",
        "Custom Visualizations", "Summary",
    ]
else:
    tab_names = [
        "General Inputs", "Data Questionnaire", "Data Input", "Asset View",
        "Unit Economics Analysis", "Churn Analysis", "LTV Analysis", "TS Covenants",
        "General Analysis", "Custom Visualizations", "Summary",
    ]
tabs = st.tabs(tab_names)

with tabs[0]:
    st.subheader("General Inputs")
    if is_lending:
        col1, col2, col3 = st.columns(3)
        col1.metric(
            "Date of extraction", str(gi.extraction_date.date()),
            help="The max date in the loan tape. Used to identify which loans have reached their maturity.",
        )
        col2.metric(
            "Days after term", gi.days_after_term,
            help="Days after term used to compute loss rate (default 90 = Term + 3 months). "
                 "Modify only if the company has significant repayments after 3 months from term.",
        )
        col3.metric(
            "Minimum loans per cohort", gi.min_loans_per_cohort,
            help="Affects the cohort stressed loss rate: requires a minimum number of observations to "
                 "include a cohort in the cohort loss rate distribution (default 10).",
        )
    else:
        col1, col2, col3 = st.columns(3)
        col1.metric("Date of extraction", str(gi.extraction_date.date()))
        col2.metric("Days after term", gi.days_after_term)
        col3.metric("Months since default", gi.months_since_default)
        st.caption("Date of extraction = MAX(start_date) of the mapped start-date column (workbook General Inputs C3).")
        col1, col2, col3 = st.columns(3)
        col1.metric("Minimum loans per cohort", gi.min_loans_per_cohort)
        col2.metric("Useful life of asset (years)", gi.useful_life_years)
        col3.metric("Status labels", f"{gi.open_label} / {gi.closed_label} / {gi.paidoff_label}")

with tabs[1]:
    st.subheader("Data Questionnaire")
    for i, q in enumerate(active_questions, 1):
        with st.expander(f"Q{i}"):
            st.write(q)

with tabs[2]:
    if is_lending:
        st.subheader("Data Input")
        st.caption(
            f"{len(raw):,} rows x {len(raw.columns)} columns uploaded. "
            f"Computed columns (Cohort, Term, Reached T+3?) used in downstream sheets."
        )
        st.dataframe(raw, width="stretch", height=400)
    else:
        st.subheader("Data Input")
        st.caption(f"{len(df):,} rows x {len(df.columns)} columns. Includes all Lendable-derived columns.")
        st.dataframe(df, width="stretch", height=400)

if is_lending:
    with tabs[3]:
        st.subheader("Cohorts")
        st.caption(f"{len(cohorts)} monthly cohorts")
        st.dataframe(cohorts, width="stretch", height=400)

    with tabs[4]:
        st.subheader(f"Cohorts for X or more loans (>= {gi.min_loans_per_cohort})")
        st.caption(f"{len(filtered)} cohorts pass the minimum-loan filter")
        st.dataframe(filtered, width="stretch", height=400)

    with tabs[5]:
        st.subheader("LTV Analysis")
        c1, c2, c3 = st.columns(3)
        c1.metric("95th Percentile Losses", fmt(ltv_data["95th Percentile Losses"]))
        c2.metric("Average Total Revenue %", fmt(ltv_data["Average Total Revenue %"]))
        c3.metric("Average Term (days)", fmt(ltv_data["Average Term"], "{:,.1f}"))
        if not lending_chart_data.empty:
            st.altair_chart(
                _add_reference_line(
                    alt.Chart(lending_chart_data).mark_line(point=True).encode(
                        x=alt.X("Cohort:T", title="Cohort"),
                        y=alt.Y("Loss Rate:Q", title="Loss Rate", axis=alt.Axis(format="%")),
                        tooltip=["Cohort:T", alt.Tooltip("Loss Rate:Q", format=".2%")],
                    ),
                    ltv_data["95th Percentile Losses"],
                    f"95th %ile: {ltv_data['95th Percentile Losses']:.2%}",
                    color="#d62728",
                    x_anchor=lending_chart_data["Cohort"].max(),
                ).properties(title="Loss per Cohort", height=350),
                width="stretch",
            )

    with tabs[6]:
        st.subheader("Unit Economics Analysis")
        row1 = st.columns(3)
        row1[0].metric("Average Expected Term (days)", fmt(ue_data["Average Expected Term"], "{:,.1f}"))
        row1[1].metric("Average Loss", fmt(ue_data["Average Loss"]))
        row1[2].metric("Loss Rate Proxy (1-PvD)", fmt(ue_data["Loss Rate Proxy (1-PvD)"]))
        row2 = st.columns(3)
        row2[0].metric("Average Principal Amount", fmt(ue_data["Average Principal Amount"], "{:,.0f}"))
        row2[1].metric("Average Fee %", fmt(ue_data["Average Fee %"]))
        row2[2].metric("Average Interest %", fmt(ue_data["Average Interest %"]))
        row3 = st.columns(3)
        row3[0].metric("Sense-check Margin", fmt(ue_data["Sense-check Margin"]))
        if not lending_chart_data.empty:
            st.altair_chart(
                _add_reference_line(
                    alt.Chart(lending_chart_data).mark_line(point=True).encode(
                        x=alt.X("Cohort:T", title="Cohort"),
                        y=alt.Y("Loss Rate:Q", title="Loss Rate", axis=alt.Axis(format="%")),
                        tooltip=["Cohort:T", alt.Tooltip("Loss Rate:Q", format=".2%")],
                    ),
                    ue_data["Average Loss"],
                    f"Avg: {ue_data['Average Loss']:.2%}",
                    color="#2ca02c",
                    x_anchor=lending_chart_data["Cohort"].max(),
                ).properties(title="Loss per Cohort", height=300),
                width="stretch",
            )
            st.altair_chart(
                _add_reference_line(
                    alt.Chart(lending_chart_data).mark_line(point=True).encode(
                        x=alt.X("Cohort:T", title="Cohort"),
                        y=alt.Y("Weighted Avg Term:Q", title="Avg Term (days)"),
                        tooltip=["Cohort:T", alt.Tooltip("Weighted Avg Term:Q", format=".1f")],
                    ),
                    ue_data["Average Expected Term"],
                    f"Avg: {ue_data['Average Expected Term']:.1f} days",
                    color="#2ca02c",
                    x_anchor=lending_chart_data["Cohort"].max(),
                ).properties(title="Average Term per Cohort (days)", height=300),
                width="stretch",
            )
            fee_int = lending_chart_data.melt(
                id_vars=["Cohort"], value_vars=["Fee %", "Interest %"],
                var_name="Metric", value_name="Pct",
            )
            st.altair_chart(
                _add_reference_line(
                    _add_reference_line(
                        alt.Chart(fee_int).mark_line(point=True).encode(
                            x=alt.X("Cohort:T", title="Cohort"),
                            y=alt.Y("Pct:Q", title="% of Principal", axis=alt.Axis(format="%")),
                            color="Metric:N",
                            tooltip=["Cohort:T", "Metric:N", alt.Tooltip("Pct:Q", format=".2%")],
                        ),
                        ue_data["Average Fee %"],
                        f"Fee avg: {ue_data['Average Fee %']:.2%}",
                        color="#1f77b4",
                        x_anchor=lending_chart_data["Cohort"].max(),
                    ),
                    ue_data["Average Interest %"],
                    f"Interest avg: {ue_data['Average Interest %']:.2%}",
                    color="#ff7f0e",
                    x_anchor=lending_chart_data["Cohort"].max(),
                ).properties(title="Average Fee and Interest Percent per Cohort", height=300),
                width="stretch",
            )

    with tabs[7]:
        st.subheader("General Analysis")
        summary = lending_general_analysis(df)
        st.write(f"**Shape:** {summary['shape'][0]:,} rows x {summary['shape'][1]} columns")
        st.write("**Columns:**", ", ".join(summary["columns"]))
        orig = cohorts[["Cohort", "Total Principal"]].dropna()
        if not orig.empty:
            st.altair_chart(
                alt.Chart(orig).mark_bar().encode(
                    x=alt.X("Cohort:T", title="Cohort"),
                    y=alt.Y("Total Principal:Q", title="Total Principal"),
                    tooltip=["Cohort:T", alt.Tooltip("Total Principal:Q", format=",.0f")],
                ).properties(title="Originations per month", height=350),
                width="stretch",
            )

    with tabs[8]:
        _render_custom_visualizations_tab({
            "Data Input (loan-level)": df, "Cohorts": cohorts, "Cohorts for X or more loans": filtered,
        })

    with tabs[9]:
        st.subheader("Summary")
        st.metric(
            "Sense-check Margin", fmt(ue_data["Sense-check Margin"]),
            help="Revenue % minus stressed loss, grossed up for that loss. The single number "
                 "that answers whether this product is economically viable after losses.",
        )
        st.markdown("---")
        s1, s2, s3, s4 = st.columns(4)
        s1.metric("Loans", f"{len(df):,}")
        s2.metric("Total Principal", f"{df['Principal Value'].sum():,.0f}")
        s3.metric("Date Range", f"{df['Disbursement Date'].min().date()} to {gi.extraction_date.date()}")
        s4.metric("Cohorts Qualifying", f"{len(filtered)} / {len(cohorts)}")
        r1, r2, r3 = st.columns(3)
        r1.metric("95th Percentile Losses", fmt(ltv_data["95th Percentile Losses"]))
        r2.metric("Average Loss", fmt(ue_data["Average Loss"]))
        r3.metric("Loss Rate Proxy (1-PvD)", fmt(ue_data["Loss Rate Proxy (1-PvD)"]))
        y1, y2, y3, y4 = st.columns(4)
        y1.metric("Average Interest %", fmt(ue_data["Average Interest %"]))
        y2.metric("Average Fee %", fmt(ue_data["Average Fee %"]))
        y3.metric("Average Total Revenue %", fmt(ltv_data["Average Total Revenue %"]))
        y4.metric("Average Term (days)", fmt(ltv_data["Average Term"], "{:,.1f}"))

else:
    with tabs[3]:
        st.subheader("Asset View")
        st.caption(f"{len(av):,} unique assets")
        st.dataframe(av, width="stretch", height=400)

    with tabs[4]:
        st.subheader("Unit Economics Analysis")
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Downpayment %", fmt(ue_data["downpayment_pct"]))
        c2.metric("Lease Tenor (m)", fmt(ue_data["lease_tenor_m"], "{:.1f}"))
        c3.metric("Margin", fmt(ue_data["margin"]))
        c4.metric("Utilisation Rate", fmt(ue_data["utilisation_rate"]))
        c1, c2, c3 = st.columns(3)
        c1.metric("PvD", fmt(ue_data["pvd"]))
        c2.metric("Historical % Collected on Principal", fmt(ue_data["historical_pct_collected_on_principal"]))
        c3.metric("Monthly Observed Repayment", fmt(ue_data["monthly_observed_repayment"], "{:.2%}"))
        curve_chart = curve.dropna(subset=["pct_avg_collection"])
        if not curve_chart.empty:
            st.altair_chart(
                alt.Chart(curve_chart).mark_line(point=True).encode(
                    x=alt.X("mob:Q", title="Month on Books"),
                    y=alt.Y("pct_avg_collection:Q", title="% Avg collection", axis=alt.Axis(format="%")),
                    tooltip=["mob", alt.Tooltip("pct_avg_collection:Q", format=".2%")],
                ).properties(title="High level repayment curve", height=350),
                width="stretch",
            )
        st.dataframe(curve, width="stretch", height=300)

    with tabs[5]:
        st.subheader("Churn Analysis")
        c1, c2, c3 = st.columns(3)
        c1.metric("95th %ile monthly churn", fmt(ca_data["pctile_95"], "{:.2%}"))
        c2.metric("Average monthly churn", fmt(ca_data["avg_churn"], "{:.2%}"))
        c3.metric("Stressed churn (1.7x)", fmt(ca_data["stress_churn"], "{:.2%}"))
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("1-year multiple", fmt(ca_data["multiplier_1y"], "{:.2f}"))
        c2.metric("2-year multiple", fmt(ca_data["multiplier_2y"], "{:.2f}"))
        c3.metric("3-year multiple", fmt(ca_data["multiplier_3y"], "{:.2f}"))
        c4.metric("Total multiple", fmt(ca_data["multiplier_total"], "{:.2f}"))
        st.altair_chart(
            alt.Chart(ca_data["residual_curve"]).mark_line().encode(
                x=alt.X("month:Q", title="Month"),
                y=alt.Y("residual_value:Q", title="Residual portfolio"),
                tooltip=["month", alt.Tooltip("residual_value:Q", format=".2%")],
            ).properties(title="Expected residual portfolio by month", height=300),
            width="stretch",
        )
        churn_chart = cohorts.dropna(subset=["churn_rate"])
        if not churn_chart.empty:
            st.altair_chart(
                alt.Chart(churn_chart).mark_line(point=True).encode(
                    x=alt.X("cohort:T", title="Cohort"),
                    y=alt.Y("churn_rate:Q", title="Churn Rate", axis=alt.Axis(format="%")),
                    tooltip=["cohort:T", alt.Tooltip("churn_rate:Q", format=".2%")],
                ).properties(title="Churn Rate per Cohort", height=300),
                width="stretch",
            )
        st.markdown("**Lease Cohorts**")
        st.dataframe(cohorts, width="stretch", height=300)
        st.markdown(f"**Cohorts for X or more loans (>= {gi.min_loans_per_cohort} new leases)**")
        st.dataframe(filtered, width="stretch", height=200)

    with tabs[6]:
        st.subheader("LTV Analysis")
        col1, col2 = st.columns(2)
        with col1:
            st.markdown("**Asset Analysis**")
            st.metric("Average Useful Life (m)", fmt(ltv_data["avg_useful_life_m"], "{:.1f}"))
            st.metric("Defaulted contracts > 3mo", fmt(ltv_data["n_defaulted_gt_3m"], "{:.0f}"))
            st.metric("% recovered", fmt(ltv_data["pct_recovered"]))
            st.metric("Loss (non-recoverability)", fmt(ltv_data["loss_non_recoverability"]))
            st.metric("MRR Multiplier (3y)", fmt(ltv_data["mrr_multiplier"], "{:.2f}"))
            st.metric("MRR / average cost", fmt(ltv_data["mrr_over_avg_cost"], "{:.2%}"))
        with col2:
            st.markdown("**Cash Flow Analysis (MRR)**")
            st.metric("MRR", fmt(ltv_data["mrr"], "{:,.0f}"))
            st.metric("Average Monthly Churn", fmt(ltv_data["avg_monthly_churn"], "{:.2%}"))
            st.metric("Average Collection Rate", fmt(ltv_data["avg_collection_rate"]))
            st.metric("95th %ile Churn", fmt(ltv_data["pctile_95_churn"], "{:.2%}"))
            st.metric("Stressed Churn", fmt(ltv_data["stressed_churn"], "{:.2%}"))

    with tabs[7]:
        st.subheader("TS Covenants")
        st.markdown("**Performance Covenants**")
        for m in (6, 12, 24):
            val = ts[f"paid_at_{m}m_over_cost"]
            suffix = fmt(val) if pd.notna(val) else "n/a (insufficient data)"
            st.write(f"Paid at {m} months over asset acquisition cost: {suffix}")
        st.metric("Option 2 - Avg Monthly Paid over Avg Cost of Asset", fmt(ts["option2_observed"], "{:.2%}"))
        st.markdown("**Recoverability Covenants**")
        st.metric("Recovery Rate (observed loss)", fmt(ts["recovery_rate_observed"]))

    with tabs[8]:
        st.subheader("General Analysis")
        summary = general_analysis(df)
        st.write(f"**Shape:** {summary['shape'][0]:,} rows x {summary['shape'][1]} columns")
        st.write("**Columns:**", ", ".join(summary["columns"]))
        orig = cohorts[["cohort", "value_of_leases"]].dropna()
        if not orig.empty:
            st.altair_chart(
                alt.Chart(orig).mark_bar().encode(
                    x=alt.X("cohort:T", title="Cohort"),
                    y=alt.Y("value_of_leases:Q", title="Value of Leases"),
                    tooltip=["cohort:T", alt.Tooltip("value_of_leases:Q", format=",.0f")],
                ).properties(title="New lease value per month", height=350),
                width="stretch",
            )

    with tabs[9]:
        _render_custom_visualizations_tab({
            "Data Input": df, "Asset View": av, "Lease Cohorts": cohorts,
            "Cohorts for X or more loans": filtered, "Repayment Curve": curve,
            "Churn Residual Curve": ca_data["residual_curve"],
        })

    with tabs[10]:
        st.subheader("Summary")
        st.metric(
            "MRR Multiplier (3y)", fmt(ltv_data["mrr_multiplier"], "{:.2f}"),
            help="Expected 3-year MRR relative to acquisition cost. The single number that "
                 "answers whether this asset-lease product is economically viable after churn "
                 "and recoveries.",
        )
        st.markdown("---")
        s1, s2, s3 = st.columns(3)
        s1.metric("Contracts", f"{len(av):,}")
        s2.metric("MRR", fmt(ltv_data["mrr"], "{:,.0f}"))
        s3.metric("Average Useful Life (m)", fmt(ltv_data["avg_useful_life_m"], "{:.1f}"))
        r1, r2, r3, r4 = st.columns(4)
        r1.metric("95th %ile Churn", fmt(ltv_data["pctile_95_churn"], "{:.2%}"))
        r2.metric("Stressed Churn", fmt(ltv_data["stressed_churn"], "{:.2%}"))
        r3.metric("Defaulted contracts > 3mo", fmt(ltv_data["n_defaulted_gt_3m"], "{:.0f}"))
        r4.metric("% recovered", fmt(ltv_data["pct_recovered"]))
        y1, y2, y3 = st.columns(3)
        y1.metric("Loss (non-recoverability)", fmt(ltv_data["loss_non_recoverability"]))
        y2.metric("MRR / average cost", fmt(ltv_data["mrr_over_avg_cost"], "{:.2%}"))
        y3.metric("Average Collection Rate", fmt(ltv_data["avg_collection_rate"]))

st.markdown("---")


def _build_lending_export_workbook():
    raw_cols = ["Loan ID", "Disbursement Date", "Expected Completion Date",
                "Principal Value", "Expected Interest", "Expected Fee",
                "Total Due", "Total Paid"]
    data_frame = pd.DataFrame({col: df.get(col, pd.Series([None] * len(df))) for col in raw_cols})
    num_data = len(df)
    orig_source = cohorts[["Cohort", "Total Principal"]].dropna()

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        gs = writer.book.create_sheet("General Inputs", 0)
        for r, label, val, note in [
            (2, "Inputs", None, None),
            (3, "Date of extraction", "=MAX('Data Input'!B:B)", "Defaults to most recent disbursement date"),
            (4, "Days after term", gi.days_after_term, "Ignore any loans for loss rates that are less than this number of days after term"),
            (5, "Minimum loans per cohort", gi.min_loans_per_cohort, "Ignore any cohorts for loss rates that are less than this number of loans"),
        ]:
            if label:
                gs.cell(row=r, column=2, value=label)
            if val is not None:
                gs.cell(row=r, column=3, value=val)
            if note:
                gs.cell(row=r, column=4, value=note)

        q_df = pd.DataFrame([(i, q) for i, q in enumerate(LENDING_QUESTIONS, 1)], columns=["#", "Question"])
        q_df.to_excel(writer, sheet_name="Data Questionnaire", index=False)

        data_frame.to_excel(writer, sheet_name="Data Input", index=False, header=True)
        dws = writer.sheets["Data Input"]
        dws.cell(row=1, column=9, value="Reached T+3?")
        dws.cell(row=1, column=10, value="Cohort")
        dws.cell(row=1, column=11, value="Term (days)")
        for r in range(2, max(num_data, 1) + 2):
            dws.cell(row=r, column=9).value = f'=IF(ISBLANK(C{r}),"",(C{r}+\'General Inputs\'!$C$4)<=\'General Inputs\'!$C$3)'
            dws.cell(row=r, column=10).value = f'=IF(ISBLANK(B{r}),"",DATE(YEAR(B{r}),MONTH(B{r}),1))'
            dws.cell(row=r, column=11).value = f'=IF(OR(ISBLANK(B{r}),ISBLANK(C{r})),"",C{r}-B{r})'

        ltv_df = pd.DataFrame(list(ltv_data.items()), columns=["Metric", "Value"])
        ltv_df.to_excel(writer, sheet_name="LTV Analysis", index=False)
        ltv_ws = writer.sheets["LTV Analysis"]
        if not lending_chart_data.empty:
            lending_chart_data[["Cohort", "Loss Rate"]].to_excel(writer, sheet_name="LTV Analysis", startrow=6, index=False)
            c = LineChart()
            c.title = "Loss per Cohort"
            c.y_axis.numFmt = "0.00%"
            c.height, c.width = 14, 24
            data = Reference(ltv_ws, min_col=1, min_row=7, max_col=2, max_row=7 + len(lending_chart_data))
            cats = Reference(ltv_ws, min_col=1, min_row=8, max_row=7 + len(lending_chart_data))
            c.add_data(data, titles_from_data=True)
            c.set_categories(cats)
            ltv_ws.add_chart(c, "B8")

        ue_df = pd.DataFrame(list(ue_data.items()), columns=["Metric", "Value"])
        ue_df.to_excel(writer, sheet_name="Unit Economics Analysis", index=False)
        ue_ws = writer.sheets["Unit Economics Analysis"]
        if not lending_chart_data.empty:
            lending_chart_data[["Cohort", "Loss Rate"]].to_excel(writer, sheet_name="Unit Economics Analysis", startrow=10, index=False)
            lending_chart_data[["Cohort", "Weighted Avg Term"]].to_excel(writer, sheet_name="Unit Economics Analysis", startrow=10, startcol=4, index=False)
            lending_chart_data[["Cohort", "Fee %", "Interest %"]].to_excel(writer, sheet_name="Unit Economics Analysis", startrow=10, startcol=8, index=False)
            sr = 11
            lr = sr + len(lending_chart_data) - 1
            for title, col, anchor in [
                ("Loss per Cohort", (1, 2), "B29"),
                ("Average Term per Cohort (days)", (5, 6), "B48"),
                ("Average Fee and Interest Percent per Cohort", (9, 11), "B67"),
            ]:
                c = LineChart()
                c.title = title
                c.height, c.width = 14, 24
                if col[1] - col[0] == 1:
                    c.y_axis.numFmt = "0.00%" if "Loss" in title else "0"
                data = Reference(ue_ws, min_col=col[0], min_row=sr - 1, max_col=col[1], max_row=lr)
                cats = Reference(ue_ws, min_col=col[0], min_row=sr, max_row=lr)
                c.add_data(data, titles_from_data=True)
                c.set_categories(cats)
                ue_ws.add_chart(c, anchor)

        ga_summary = lending_general_analysis(df)
        ga_df = pd.DataFrame([
            ("Shape", f"{ga_summary['shape'][0]:,} rows x {ga_summary['shape'][1]} cols"),
            ("Columns", ", ".join(ga_summary["columns"])),
        ], columns=["Property", "Value"])
        ga_df.to_excel(writer, sheet_name="General Analysis", index=False)
        ga_ws = writer.sheets["General Analysis"]
        if not orig_source.empty:
            orig_source.to_excel(writer, sheet_name="General Analysis", startrow=4, index=False)
            c = BarChart()
            c.title = "Originations per month"
            c.height, c.width = 14, 24
            data = Reference(ga_ws, min_col=1, min_row=5, max_col=2, max_row=4 + len(orig_source))
            cats = Reference(ga_ws, min_col=1, min_row=6, max_row=4 + len(orig_source))
            c.add_data(data, titles_from_data=True)
            c.set_categories(cats)
            ga_ws.add_chart(c, "A6")

        _write_custom_charts_sheet(writer, st.session_state.get("export_charts", []))

        cohorts.to_excel(writer, sheet_name="Cohorts", index=False)
        filtered.to_excel(writer, sheet_name="Cohorts for X or more loans", index=False)

    buf.seek(0)
    return buf


def _build_export_workbook():
    export_col_order = REQUIRED_COLUMNS + OPTIONAL_COLUMNS
    data_input_export = df[export_col_order + [c for c in df.columns if c not in export_col_order]]

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        gs = writer.book.create_sheet("General Inputs", 0)
        # start_date is the 3rd column (C) in the fixed Data Input export order below.
        for r, label, val, note in [
            (2, "Inputs", None, None),
            (3, "Date of extraction", "=MAX('Data Input'!C:C)", "Defaults to most recent contract start date"),
            (4, "Days after term", gi.days_after_term, "Ignore any contracts for loss rates that are less than this number of days after term"),
            (5, "Months since default", gi.months_since_default, "Includes only contracts that defaulted over this many months ago"),
            (6, "Minimum loans per cohort", gi.min_loans_per_cohort, "Ignore any cohorts for loss rates that are less than this number of contracts"),
            (7, "Useful life of asset (years)", gi.useful_life_years, "Assumed Useful Life of Asset"),
        ]:
            if label:
                gs.cell(row=r, column=2, value=label)
            if val is not None:
                gs.cell(row=r, column=3, value=val)
            if note:
                gs.cell(row=r, column=4, value=note)

        q_df = pd.DataFrame([(i, q) for i, q in enumerate(QUESTIONS, 1)], columns=["#", "Question"])
        q_df.to_excel(writer, sheet_name="Data Questionnaire", index=False)

        data_input_export.to_excel(writer, sheet_name="Data Input", index=False, header=True)

        av.to_excel(writer, sheet_name="Asset View", index=False)

        ue_df = pd.DataFrame(list(ue_data.items()), columns=["Metric", "Value"])
        ue_df = ue_df[ue_df["Metric"] != "residual_curve"]
        ue_df.to_excel(writer, sheet_name="Unit Economics Analysis", index=False)
        ue_ws = writer.sheets["Unit Economics Analysis"]
        curve_export = curve.dropna(subset=["pct_avg_collection"])
        if not curve_export.empty:
            curve_export[["mob", "pct_avg_collection"]].to_excel(
                writer, sheet_name="Unit Economics Analysis", startrow=len(ue_df) + 2, index=False,
            )
            sr = len(ue_df) + 3
            lr = sr + len(curve_export) - 1
            c = LineChart()
            c.title = "High level repayment curve"
            c.y_axis.numFmt = "0.00%"
            c.height, c.width = 14, 24
            data = Reference(ue_ws, min_col=1, min_row=sr - 1, max_col=2, max_row=lr)
            cats = Reference(ue_ws, min_col=1, min_row=sr, max_row=lr)
            c.add_data(data, titles_from_data=True)
            c.set_categories(cats)
            ue_ws.add_chart(c, f"D{sr - 1}")

        ca_export = {k: v for k, v in ca_data.items() if k != "residual_curve"}
        ca_df = pd.DataFrame(list(ca_export.items()), columns=["Metric", "Value"])
        ca_df.to_excel(writer, sheet_name="Churn Analysis", index=False)
        ca_ws = writer.sheets["Churn Analysis"]
        residual = ca_data["residual_curve"]
        residual.to_excel(writer, sheet_name="Churn Analysis", startrow=len(ca_df) + 2, index=False)
        sr = len(ca_df) + 3
        lr = sr + len(residual) - 1
        c = LineChart()
        c.title = "Expected residual portfolio by month"
        c.height, c.width = 14, 24
        data = Reference(ca_ws, min_col=1, min_row=sr - 1, max_col=2, max_row=lr)
        cats = Reference(ca_ws, min_col=1, min_row=sr, max_row=lr)
        c.add_data(data, titles_from_data=True)
        c.set_categories(cats)
        ca_ws.add_chart(c, f"D{sr - 1}")

        ltv_df = pd.DataFrame(list(ltv_data.items()), columns=["Metric", "Value"])
        ltv_df.to_excel(writer, sheet_name="LTV Analysis", index=False)

        ts_df = pd.DataFrame(list(ts.items()), columns=["Metric", "Value"])
        ts_df.to_excel(writer, sheet_name="TS Covenants", index=False)

        ga_summary = general_analysis(df)
        ga_df = pd.DataFrame([
            ("Shape", f"{ga_summary['shape'][0]:,} rows x {ga_summary['shape'][1]} cols"),
            ("Columns", ", ".join(ga_summary["columns"])),
        ], columns=["Property", "Value"])
        ga_df.to_excel(writer, sheet_name="General Analysis", index=False)
        ga_ws = writer.sheets["General Analysis"]
        orig_source = cohorts[["cohort", "value_of_leases"]].dropna()
        if not orig_source.empty:
            orig_source.to_excel(writer, sheet_name="General Analysis", startrow=4, index=False)
            c = BarChart()
            c.title = "New lease value per month"
            c.height, c.width = 14, 24
            data = Reference(ga_ws, min_col=1, min_row=5, max_col=2, max_row=4 + len(orig_source))
            cats = Reference(ga_ws, min_col=1, min_row=6, max_row=4 + len(orig_source))
            c.add_data(data, titles_from_data=True)
            c.set_categories(cats)
            ga_ws.add_chart(c, "D6")

        cohorts.to_excel(writer, sheet_name="Lease Cohorts", index=False)
        filtered.to_excel(writer, sheet_name="Cohorts for X or more loans", index=False)

        _write_custom_charts_sheet(writer, st.session_state.get("export_charts", []))

    buf.seek(0)
    return buf


if is_lending:
    buf = _build_lending_export_workbook()
    file_name = f"SC_Analysis_Lending_{pd.Timestamp.now():%Y-%m-%d}.xlsx"
else:
    buf = _build_export_workbook()
    file_name = f"SC_Analysis_Rental_{pd.Timestamp.now():%Y-%m-%d}.xlsx"

st.download_button(
    "Download full workbook as Excel",
    buf,
    file_name=file_name,
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)
